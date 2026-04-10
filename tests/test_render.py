from __future__ import annotations

import json
import os
import sys
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image
from typer.testing import CliRunner

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from herndon.cli import app
from herndon.utils import dump_json


runner = CliRunner()


def test_render_workbook_with_theme_and_table(tmp_path: Path) -> None:
    runner.invoke(app, ["init", str(tmp_path)])
    cwd = Path.cwd()
    try:
        os.chdir(tmp_path)
        runner.invoke(app, ["new", "workbook", "finance"])
        workbook_path = tmp_path / "workbooks" / "finance" / "workbook.json"
        runner.invoke(app, ["new", "sheet", str(workbook_path), "summary"])
        dump_json(
            tmp_path / ".herndon" / "themes" / "brand.json",
            {
                "name": "brand",
                "colors": {"text": "#111111", "header_fill": "#EEEEEE"},
                "fonts": {"default": "Aptos"},
                "styles": {
                    "header": {"bold": True, "fill": "#EEEEEE"},
                    "currency": {"number_format": "#,##0.00"},
                },
            },
        )
        workbook = json.loads(workbook_path.read_text())
        workbook["theme"] = "brand"
        workbook_path.write_text(json.dumps(workbook, indent=2))
        runner.invoke(
            app,
            [
                "sheets",
                "set-range",
                str(workbook_path),
                "summary",
                "A1",
                "--data-json",
                '[["Quarter","Revenue"],["Q1",100],["Q2",150]]',
                "--row-styles",
                '{"0":"header"}',
                "--col-styles",
                '{"1":"currency"}',
            ],
        )
        runner.invoke(app, ["sheets", "add-table", str(workbook_path), "summary", "revenue_table", "--ref", "A1:B3", "--name", "RevenueData"])
        result = runner.invoke(app, ["render", str(workbook_path), "--format", "json"])
        assert result.exit_code == 0
        output_path = tmp_path / ".herndon" / "builds" / "finance" / "finance.xlsx"
        assert output_path.exists()
        wb = load_workbook(output_path)
        ws = wb["Summary"]
        assert ws["A2"].value == "Q1"
        assert ws["B3"].value == 150
        manifest = json.loads((tmp_path / ".herndon" / "builds" / "finance" / "manifest.json").read_text())
        assert manifest["workbook_id"] == "finance"
    finally:
        os.chdir(cwd)


def test_render_workbook_with_pie_chart(tmp_path: Path) -> None:
    runner.invoke(app, ["init", str(tmp_path)])
    cwd = Path.cwd()
    try:
        os.chdir(tmp_path)
        runner.invoke(app, ["new", "workbook", "expenses"])
        workbook_path = tmp_path / "workbooks" / "expenses" / "workbook.json"
        runner.invoke(app, ["new", "sheet", str(workbook_path), "overview"])
        dump_json(
            tmp_path / ".herndon" / "themes" / "brand.json",
            {
                "name": "brand",
                "styles": {
                    "header": {"bold": True},
                    "currency": {"number_format": "$#,##0.00"},
                },
            },
        )
        workbook = json.loads(workbook_path.read_text())
        workbook["theme"] = "brand"
        workbook_path.write_text(json.dumps(workbook, indent=2))
        sheet_path = tmp_path / "workbooks" / "expenses" / "sheets" / "001-overview.json"
        dump_json(
            sheet_path,
            {
                "sheet_id": "overview",
                "title": "Overview",
                "freeze_rows": 1,
                "freeze_cols": 0,
                "zoom": 100,
                "column_widths": {},
                "row_heights": {},
                "cells": [],
                "ranges": [
                    {
                        "anchor": "A1",
                        "data": [["Category", "Amount"], ["Rent", 1000], ["Travel", 200], ["Software", 300]],
                        "row_styles": {"0": "header"},
                        "col_styles": {"1": "currency"},
                    }
                ],
                "merges": [],
                "tables": [],
                "charts": [
                    {
                        "chart_id": "expense_pie",
                        "chart_type": "pie",
                        "title": "Spend by Category",
                        "anchor": "D2",
                        "w": 6,
                        "h": 4,
                        "series": [
                            {
                                "label": "Amount",
                                "values": "'Overview'!$B$2:$B$4",
                                "categories": "'Overview'!$A$2:$A$4",
                            }
                        ],
                        "show_legend": True,
                    }
                ],
            },
        )
        result = runner.invoke(app, ["render", str(workbook_path), "--format", "json"])
        assert result.exit_code == 0
        output_path = tmp_path / ".herndon" / "builds" / "expenses" / "expenses.xlsx"
        wb = load_workbook(output_path)
        ws = wb["Overview"]
        assert ws._charts
        assert ws._charts[0].title is not None
    finally:
        os.chdir(cwd)


def test_render_workbook_with_stacked_bar_and_line_charts(tmp_path: Path) -> None:
    runner.invoke(app, ["init", str(tmp_path)])
    cwd = Path.cwd()
    try:
        os.chdir(tmp_path)
        runner.invoke(app, ["new", "workbook", "charts"])
        workbook_path = tmp_path / "workbooks" / "charts" / "workbook.json"
        workbook = json.loads(workbook_path.read_text())
        workbook["sheets"] = ["sheets/001-stacked.json", "sheets/002-line.json"]
        workbook_path.write_text(json.dumps(workbook, indent=2))

        dump_json(
            tmp_path / "workbooks" / "charts" / "sheets" / "001-stacked.json",
            {
                "sheet_id": "stacked",
                "title": "Stacked",
                "freeze_rows": 1,
                "freeze_cols": 0,
                "zoom": 100,
                "column_widths": {},
                "row_heights": {},
                "cells": [],
                "ranges": [
                    {
                        "anchor": "A1",
                        "data": [
                            ["Month", "Rent", "Payroll", "Software"],
                            ["Jan", 2400, 8500, 1200],
                            ["Feb", 2400, 8500, 1250],
                            ["Mar", 2400, 8500, 1300],
                        ],
                    }
                ],
                "merges": [],
                "tables": [],
                "charts": [
                    {
                        "chart_id": "stacked_bar",
                        "chart_type": "bar",
                        "title": "Monthly Spend Mix",
                        "anchor": "F2",
                        "w": 8,
                        "h": 5,
                        "percent_stacked": True,
                        "legend_position": "b",
                        "show_data_labels": True,
                        "x_axis_title": "Spend Share",
                        "y_axis_title": "Month",
                        "series": [
                            {"label": "Rent", "values": "'Stacked'!$B$2:$B$4", "categories": "'Stacked'!$A$2:$A$4", "color": "#2563EB"},
                            {"label": "Payroll", "values": "'Stacked'!$C$2:$C$4", "categories": "'Stacked'!$A$2:$A$4", "color": "#059669"},
                            {"label": "Software", "values": "'Stacked'!$D$2:$D$4", "categories": "'Stacked'!$A$2:$A$4", "color": "#D97706"},
                        ],
                        "show_legend": True,
                    }
                ],
            },
        )

        dump_json(
            tmp_path / "workbooks" / "charts" / "sheets" / "002-line.json",
            {
                "sheet_id": "line",
                "title": "Line",
                "freeze_rows": 1,
                "freeze_cols": 0,
                "zoom": 100,
                "column_widths": {},
                "row_heights": {},
                "cells": [],
                "ranges": [
                    {
                        "anchor": "A1",
                        "data": [
                            ["Month", "Total"],
                            ["Jan", 12100],
                            ["Feb", 12150],
                            ["Mar", 12200],
                        ],
                    }
                ],
                "merges": [],
                "tables": [],
                "charts": [
                    {
                        "chart_id": "spend_line",
                        "chart_type": "line",
                        "title": "Total Spend Trend",
                        "anchor": "D2",
                        "w": 7,
                        "h": 4,
                        "legend_position": "r",
                        "y_axis_title": "Spend",
                        "series": [
                            {"label": "Total", "values": "'Line'!$B$2:$B$4", "categories": "'Line'!$A$2:$A$4", "color": "#7C3AED"}
                        ],
                        "show_legend": True,
                    }
                ],
            },
        )

        result = runner.invoke(app, ["render", str(workbook_path), "--format", "json"])
        assert result.exit_code == 0
        output_path = tmp_path / ".herndon" / "builds" / "charts" / "charts.xlsx"
        wb = load_workbook(output_path)
        assert wb["Stacked"]._charts
        assert wb["Line"]._charts
        assert wb["Stacked"]._charts[0].grouping == "percentStacked"
        assert wb["Stacked"]._charts[0].legend.position == "b"
        assert wb["Line"]._charts[0].legend.position == "r"
        assert wb["Stacked"]._charts[0].series[0].tx.v == "Rent"
    finally:
        os.chdir(cwd)


def test_render_workbook_with_embedded_image_and_asset_inspect(tmp_path: Path) -> None:
    runner.invoke(app, ["init", str(tmp_path)])
    cwd = Path.cwd()
    try:
        os.chdir(tmp_path)
        runner.invoke(app, ["new", "workbook", "assets_demo"])
        workbook_path = tmp_path / "workbooks" / "assets_demo" / "workbook.json"
        runner.invoke(app, ["new", "sheet", str(workbook_path), "overview"])
        asset_path = tmp_path / "assets" / "logo.png"
        Image.new("RGB", (120, 40), color=(15, 118, 110)).save(asset_path)

        inspect_result = runner.invoke(app, ["assets", "inspect", str(asset_path), "--format", "json"])
        assert inspect_result.exit_code == 0
        inspect_payload = json.loads(inspect_result.stdout)
        assert inspect_payload["width"] == 120
        assert inspect_payload["height"] == 40

        sheet_path = tmp_path / "workbooks" / "assets_demo" / "sheets" / "001-overview.json"
        dump_json(
            sheet_path,
            {
                "sheet_id": "overview",
                "title": "Overview",
                "freeze_rows": 0,
                "freeze_cols": 0,
                "zoom": 100,
                "column_widths": {},
                "row_heights": {},
                "cells": [
                    {
                        "cell": "B2",
                        "image_path": "assets/logo.png",
                        "w": 2.5,
                        "h": 0.8
                    }
                ],
                "ranges": [],
                "merges": [],
                "tables": [],
                "charts": []
            },
        )

        result = runner.invoke(app, ["render", str(workbook_path), "--format", "json"])
        assert result.exit_code == 0
        output_path = tmp_path / ".herndon" / "builds" / "assets_demo" / "assets_demo.xlsx"
        wb = load_workbook(output_path)
        ws = wb["Overview"]
        assert len(ws._images) == 1
    finally:
        os.chdir(cwd)
