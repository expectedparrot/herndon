from __future__ import annotations

from datetime import date
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from .errors import HerndonError
from .models import LoadedWorkbook, ThemeStyle
from .utils import looks_like_iso_date, normalize_excel_color, timestamp_utc
from .validation import validate_workbook


def _style_to_openpyxl(style: ThemeStyle):
    font = Font(
        name=style.font_name,
        size=style.font_size,
        bold=style.bold,
        italic=style.italic,
        underline="single" if style.underline else None,
        color=normalize_excel_color(style.color) if style.color else None,
    )
    fill = PatternFill(fill_type="solid", fgColor=normalize_excel_color(style.fill)) if style.fill else PatternFill()
    border_color = normalize_excel_color(style.border_color) if style.border_color else "000000"

    def border_side(name: str | None):
        return Side(style=name, color=border_color) if name else Side()

    border = Border(
        top=border_side(style.border_top),
        bottom=border_side(style.border_bottom),
        left=border_side(style.border_left),
        right=border_side(style.border_right),
    )
    alignment = Alignment(
        horizontal=style.alignment,
        vertical={"middle": "center"}.get(style.vertical_alignment, style.vertical_alignment),
        wrap_text=style.wrap_text,
    )
    return font, fill, border, alignment


def _apply_style(cell, style_name: str | None, styles: dict[str, ThemeStyle]) -> None:
    if not style_name:
        return
    style = styles[style_name]
    font, fill, border, alignment = _style_to_openpyxl(style)
    cell.font = font
    cell.fill = fill
    cell.border = border
    cell.alignment = alignment
    if style.number_format:
        cell.number_format = style.number_format


def _write_value(cell, value):
    if looks_like_iso_date(value):
        cell.value = date.fromisoformat(value)
        if cell.number_format == "General":
            cell.number_format = "YYYY-MM-DD"
    else:
        cell.value = value


def _chart_for_type(chart_type: str):
    if chart_type == "bar":
        return BarChart()
    if chart_type == "column":
        chart = BarChart()
        chart.type = "col"
        return chart
    if chart_type == "line":
        return LineChart()
    if chart_type == "pie":
        return PieChart()
    if chart_type == "scatter":
        return ScatterChart()
    raise HerndonError("render_error", f"Unsupported chart type '{chart_type}'")


def render_workbook(loaded: LoadedWorkbook) -> dict[str, str | int]:
    result = validate_workbook(loaded)
    if not result["ok"]:
        errors = [item for item in result["issues"] if item["severity"] == "error"]
        raise HerndonError("validation_error", f"Workbook contains {len(errors)} validation errors", details=result["issues"])

    styles = loaded.theme.styles if loaded.theme else {}
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet in loaded.sheets:
        ws = workbook.create_sheet(title=sheet.spec.title)
        if sheet.spec.tab_color:
            ws.sheet_properties.tabColor = normalize_excel_color(sheet.spec.tab_color)
        ws.sheet_view.zoomScale = sheet.spec.zoom

        for column, width in sheet.spec.column_widths.items():
            ws.column_dimensions[column].width = width
        for row, height in sheet.spec.row_heights.items():
            ws.row_dimensions[int(row)].height = height

        for range_spec in sheet.spec.ranges:
            start_col_letter, start_row = coordinate_from_string(range_spec.anchor)
            start_col = column_index_from_string(start_col_letter)
            for row_offset, row_values in enumerate(range_spec.data):
                for col_offset, value in enumerate(row_values):
                    cell = ws.cell(row=start_row + row_offset, column=start_col + col_offset)
                    _write_value(cell, value)
                    _apply_style(cell, range_spec.row_styles.get(str(row_offset)), styles)
                    _apply_style(cell, range_spec.col_styles.get(str(col_offset)), styles)

        for cell_spec in sheet.spec.cells:
            cell = ws[cell_spec.cell]
            if cell_spec.formula is not None:
                cell.value = cell_spec.formula
            elif cell_spec.image_path is None:
                _write_value(cell, cell_spec.value)
            _apply_style(cell, cell_spec.style, styles)
            if cell_spec.image_path:
                image_path = Path(cell_spec.image_path)
                if not image_path.is_absolute():
                    image_path = loaded.project_root / image_path
                image = XLImage(str(image_path))
                image.width = cell_spec.w * 96
                image.height = cell_spec.h * 96
                ws.add_image(image, cell_spec.cell)

        for merge in sheet.spec.merges:
            ws.merge_cells(merge)

        for table_spec in sheet.spec.tables:
            table = Table(displayName=table_spec.name, ref=table_spec.ref)
            if table_spec.style:
                table.tableStyleInfo = TableStyleInfo(
                    name=table_spec.style,
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
            ws.add_table(table)

        for chart_spec in sheet.spec.charts:
            chart = _chart_for_type(chart_spec.chart_type)
            chart.title = chart_spec.title
            chart.width = chart_spec.w
            chart.height = chart_spec.h
            if chart_spec.stacked and chart_spec.chart_type in {"bar", "column"}:
                chart.grouping = "stacked"
                chart.overlap = 100
            if chart_spec.percent_stacked and chart_spec.chart_type in {"bar", "column"}:
                chart.grouping = "percentStacked"
                chart.overlap = 100
            chart.legend = None if not chart_spec.show_legend else chart.legend
            if chart.legend is not None and chart_spec.legend_position:
                chart.legend.position = chart_spec.legend_position
            if chart_spec.show_data_labels or chart_spec.show_percent_labels:
                chart.dLbls = DataLabelList()
                chart.dLbls.showVal = chart_spec.show_data_labels
                chart.dLbls.showPercent = chart_spec.show_percent_labels
            if getattr(chart, "x_axis", None) and chart_spec.x_axis_title:
                chart.x_axis.title = chart_spec.x_axis_title
            if getattr(chart, "y_axis", None) and chart_spec.y_axis_title:
                chart.y_axis.title = chart_spec.y_axis_title
            for series_spec in chart_spec.series:
                values = Reference(range_string=series_spec.values)
                chart.add_data(values, titles_from_data=False)
                series = chart.series[-1]
                series.tx = SeriesLabel(v=series_spec.label)
                if series_spec.color:
                    color = normalize_excel_color(series_spec.color)
                    series.graphicalProperties.solidFill = color
                    if getattr(series.graphicalProperties, "line", None):
                        series.graphicalProperties.line.solidFill = color
                if series_spec.categories:
                    categories = Reference(range_string=series_spec.categories)
                    chart.set_categories(categories)
            if chart_spec.value_format and getattr(chart, "y_axis", None):
                chart.y_axis.numFmt = chart_spec.value_format
            ws.add_chart(chart, chart_spec.anchor)

        if sheet.spec.freeze_rows or sheet.spec.freeze_cols:
            ws.freeze_panes = f"{get_column_letter(sheet.spec.freeze_cols + 1)}{sheet.spec.freeze_rows + 1}"

    output_path = loaded.project_root / loaded.spec.build.output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile(suffix=".xlsx", dir=output_path.parent, delete=False) as handle:
        temp_path = Path(handle.name)
    workbook.save(temp_path)
    temp_path.replace(output_path)

    manifest_path = output_path.parent / "manifest.json"
    manifest = {
        "workbook_id": loaded.spec.workbook_id,
        "source_path": str(loaded.path.relative_to(loaded.project_root)),
        "output_path": str(output_path.relative_to(loaded.project_root)),
        "rendered_at": timestamp_utc(),
        "sheet_count": len(loaded.sheets),
        "theme": loaded.spec.theme,
    }
    manifest_path.write_text(__import__("json").dumps(manifest, indent=2) + "\n", encoding="utf-8")
    return manifest
